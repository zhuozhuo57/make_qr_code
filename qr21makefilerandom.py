# -*- coding:utf-8 -*-
import configparser
import os
import random

import openpyxl
from PIL import ImageFont

import qrcode

# '''实现随机生成参数 生成QR码组合字符'''
# version_value = random.randrange(1, 2, 1)
# ERROR_CORRECT_random_vlue = random.randrange(1, 4, 1)
# print("ERROR_CORRECT_random_vlue is", ERROR_CORRECT_random_vlue)
# box_size_value = random.randrange(1, 2, 1)
# print("box_size_value is", box_size_value)
# border_value = random.randrange(99, 100, 1)
# print("version_value is", version_value)
# qr = qrcode.QRCode(version=21, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=box_size_value,
#                    border=border_value, )
'''实现xlsx 文件路径'''
file = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
print(config)
s = config.read(file + r'\config.ini')
print(s)
excelPath = config.get("QRLocal", "QRResultPath")
filePath = config.get("QRLocal", "QRlogPath")
data = openpyxl.load_workbook(excelPath)
sheet = data['Sheet1']
''' #码制	Version版本号	内容	密度	长度	密度mm	长度mm	浓度'''
sheet.cell(1, 1).value = '码制qr'
sheet.cell(1, 2).value = "Version版本号"
sheet.cell(1, 3).value = "内容add_data"
sheet.cell(1, 4).value = "容错率ERROR_CORRECT"
sheet.cell(1, 5).value = "像素尺寸box_size"  # box_size表示每个模块的像素数，“每边模块数” × box_size 就是二维码的像素尺寸了。公式如下：
# (21 + (version - 1) * 4 + border * 2) * box_size
sheet.cell(1, 6).value = "每边模块数量border_value"  # 边距是以模块数量为单位，比如border=4就是表示4个模块。
# sheet.cell(1, 7).value = "浓度Concentration"#浓度  70% 50% 30%.
data.save(excelPath)
# pixel_size
''''''

s1 = r'纯中文：名称来历随着互联网技术的迅猛发展，中小学生、教师、家长对作文的需求越发显得多元化，随之而出现了“作文网”。诸多作文类网站推动了中国作文教育，切实方便了孩子们和语文老师。真真正正解决作文教学问题，让孩子们“有效提分，健康成长”——是作文类网站应该探索的课题。作文吧来源于学生们讨论作文的交流贴吧，聚集众多学生以及老师，以吧主吧友的模式紧密联系学生、老师、家长，用作文写出生活，用作文诠释成长，用作文培养未来的花朵。'  # 2014年，作文吧脱胎换骨，版面焕然一新，从作文素材、作文理论、作文方法、作文写作指导、作文点评等多个方面为学生、老师、家长提供更加优质的服务。发展历程1.2005年作文吧上线。2.2009年作文网第3版上线，丰富了栏目内容，提供简洁的作文指导服务。3.2012年作文吧栏目调整，专注细分年级分类作文辅导，按照年级做具有针对性的作文点评和辅导。'#4.作文吧成功改版提供更全服务，精简了部分栏目，主打小初中作文辅导与作文素材分享。'

s2 = r'English：Acting as a symbol of hope for refugees worldwide and bringing global attention to the refugee crisis, the athletes took part in the Olympic Games Rio 2016, marching and competing under the Olympic flag.Swimming had always been a passion for the family as father Ezzat is a swimming instructor dedicating his life to water.'  # He taught his three-year-old daughter to swim.That boat ride was supposed to last 45 minutes. It was just a 10km ride. The boat, meant for six to seven people, was already broken when 20 people boarded.'#Twenty minutes in, Mardini found herself, her sister, a friend of her father’s and two others in the water, pushing the broken boat ashore after more than three hours.'

s3 = r'中英文混合：International Olympic Chief Thomas Bach declared the Tokyo Games "unprecedented" and the "most challenging Olympic journey" as he addressed the near-empty 68,000-seat Olympic Stadium at the closing ceremony.闭幕式在原本可容纳6.8万人的奥运主体育场空场举行，国际奥委会主席托马斯•巴赫在致辞中说，东京奥运会是一届前所未有的奥运会，是一趟最具挑战的奥运旅程。'  # "In these difficult times we are all living through, you give the world the most precious of gifts: hope," Bach told athletes attending the ceremony.巴赫向在场的运动员们表示："在困难时期，你们带给了全世界最珍贵的礼物，希望。"Over the 16-day sports extravaganza in Tokyo, a total of 339 medal events were held across 33 sports with skateboarding, surfing, sport climbing and karate making their debuts.东京奥运会历时16天，共有33个大项，最终产生339枚金牌，新设滑板、冲浪、竞技攀岩和空手道等项目。'#A total of 94 different countries and regions claimed a medal at Tokyo, more than at any other Games.共有94个国家或地区的运动员获得东京奥运会奖牌，创历届奥运会之最。The US topped the medal table with 39 golds, 41 silvers and 33 bronzes. Team China was in second with 38 gold, 32 silver and 18 bronze medals, followed by Japan with 27 gold, 14 silver and 17 bronze medals.美国以39金41银33铜位列东京奥运奖牌榜第一，中国以38金32银18铜位列奖牌榜第二，东道主日本以27金14银17铜位居第三。'

s4 = r'标点符号：❤❥웃유♋☮✌☏☢☠✔☑♚▲♪✈✞÷↑↓◆◇⊙■□△▽¿─│♥❣♂♀☿Ⓐ✍✉☣☤✘☒♛▼♫⌘☪≈←→◈◎☉★☆⊿※¡━┃♡ღツ☼☁❅♒✎©®™Σ✪✯☭➳卐√↖↗●◐Θ◤◥︻〖〗┄┆℃℉°✿ϟ☃☂✄¢€£∞✫★½✡×↙↘○◑⊕◣◢︼【】┅┇☽☾✚〓▂▃▄▅▆▇█▉▊▋▌▍▎▏↔↕☽☾の•▸◂▴▾┈┊①②③④⑤⑥⑦⑧⑨⑩ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ㍿▓♨♛❖♓☪✙┉┋☹☺☻تヅツッシÜϡﭢ™℠℗©®♥❤❥❣❦❧♡۵웃유ღ♋♂♀☿☼☀☁☂☄☾☽❄☃☈⊙☉℃℉❅✺ϟ☇♤♧♡♢♠♣♥♦☜☞☝✍☚☛☟✌✽✾✿❁❃❋❀⚘☑✓✔√☐,'  # ☒✗✘ㄨ✕✖✖⋆✢✣✤✥❋✦✧✩✰✪✫✬✭✮✯❂✡★✱✲✳✴✵✶✷✸✹✺✻✼❄❅❆❇❈❉❊†☨✞✝☥☦☓☩☯☧☬☸✡♁✙♆。，、＇：∶；?‘’“”〝〞ˆˇ﹕︰﹔﹖﹑•¨….¸;！´？！～—ˉ｜‖＂〃｀@﹫¡¿﹏﹋﹌︴々﹟#﹩$﹠&﹪%*﹡﹢﹦﹤‐￣¯―﹨ˆ˜﹍﹎+=<＿_-\ˇ~﹉﹊（）〈〉‹›﹛﹜'#『』〖〗［］《》〔〕{}「」【】︵︷︿︹︽_﹁﹃︻︶︸﹀︺︾ˉ﹂﹄︼☩☨☦✞✛✜✝✙✠✚†‡◉○◌◍◎●◐◑◒◓◔◕◖◗❂☢⊗⊙◘◙◍⅟½⅓⅕⅙⅛⅔⅖⅚⅜¾⅗⅝⅞⅘≂≃≄≅≆≇≈≉≊≋≌≍≎≏≐≑≒≓≔≕≖≗≘≙≚≛≜≝≞≟≠≡≢≣≤≥≦≧≨≩⊰⊱⋛⋚∫∬∭∮∯∰∱∲∳%℅‰‱㊣㊎㊍㊌㊋㊏.'#㊐㊊㊚㊛㊤㊥㊦㊧㊨㊒㊞㊑㊒㊓㊔㊕㊖㊗㊘㊜㊝㊟㊠㊡㊢㊩㊪㊫㊬㊭㊮㊯㊰㊙㉿囍♔♕♖♗♘♙♚♛♜♝♞♟ℂℍℕℙℚℝℤℬℰℯℱℊℋℎℐℒℓℳℴ℘ℛℭ℮ℌℑℜℨ♪♫♩♬♭♮♯°øⒶ☮✌☪✡☭✯卐✐✎✏✑✒✍✉✁✂✃✄✆✉☎☏➟➡➢➣➤➥➦➧➨➚➘➙➛➜➝➞➸♐➲➳⏎➴➵➶➷➸➹➺➻➼➽←↑→↓↔↕↖↗↘↙↚↛↜↝↞↟↠↡↢↣↤↥↦↧↨➫➬➩➪➭➮➯➱↩↪↫↬↭↮↯↰↱↲↳↴↵↶↷↸↹↺↻↼↽↾↿⇀⇁⇂⇃⇄⇅⇆⇇⇈⇉⇊⇋⇌⇍⇎⇏⇐⇑⇒⇓⇔⇕⇖⇗⇘⇙⇚⇛⇜⇝⇞⇟⇠⇡⇢⇣⇤⇥⇦⇧⇨⇩⇪➀➁➂➃➄➅➆➇➈➉➊➋➌➍➎➏➐➑➒➓㊀㊁㊂㊃㊄㊅㊆㊇㊈㊉ⒶⒷⒸⒹⒺⒻⒼⒽⒾⒿⓀⓁⓂⓃⓄⓅⓆⓇⓈⓉⓊⓋⓌⓍⓎⓏⓐⓑⓒⓓⓔⓕⓖⓗⓘⓙⓚⓛⓜⓝⓞⓟⓠⓡⓢⓣⓤⓥⓦⓧⓨⓩ⒜⒝⒞⒟⒠⒡⒢⒣⒤⒥⒦⒧⒨⒩⒪⒫⒬⒭⒮⒯⒰⒱⒲⒳⒴⒵ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩⅪⅫⅬⅭⅮⅯⅰⅱⅲⅳⅴⅵⅶⅷⅸⅹⅺⅻⅼⅽⅾⅿ┌┍┎┏┐┑┒┓└┕┖┗┘┙┚┛├┝┞┟┠┡┢┣┤┥┦┧┨┩┪┫┬┭┮┯┰┱┲┳┴┵┶┷┸┹┺┻┼┽┾┿╀╁╂╃╄╅╆╇╈╉╊╋╌╍╎╏═║╒╓╔╕╖╗╘╙╚╛╜╝╞╟╠╡╢╣╤╥╦╧╨╩╪╫╬◤◥◄►▶◀◣◢▲▼◥▸◂▴▾△▽▷◁⊿▻◅▵▿▹◃❏❐❑❒▀▁▂▃▄▅▆▇▉▊▋█▌▍▎▏▐░▒▓▔▕■□▢▣▤▥▦▧▨▩▪▫▬▭▮▯㋀㋁㋂㋃㋄㋅㋆㋇㋈㋉㋊㋋㏠㏡㏢㏣㏤㏥㏦㏧㏨㏩㏪㏫㏬㏭㏮㏯㏰㏱㏲㏳㏴㏵㏶㏷㏸㏹㏺㏻㏼㏽㏾'

# '''实现list中读取内容生成QR码保存图片'''
# data = ['1111111', 'abcder', 'ABCDabcd','!@#$%^&*(){}_+|~:"<>?']
data_values = [s1, s2, s3, s4]
version_values = [21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40]
error_corrections = [1, 0, 3, 2]
# error_corrections = [1, 0, 3, 2]
# box_size_values = [10, 20, 30, 40, 50]
# border_values = [1, 3, 5, 7, 9]

j = 0
for version_value in version_values:
    version_value = version_values[j]
    # j += 1
    print("version_value is", version_value)

    # k = 0
    # for error_correction_value in error_corrections:
    #     error_correction_value = error_corrections[k]
    #     k += 1
    #     print("error_correction is", error_correction_value)
    #     l = 0
    #     for box_size_value in box_size_values:
    #         box_size_value = box_size_values[l]
    #         l += 1
    #         print("box_size_value is", box_size_value)
    #         m = 0
    #         for border_value in border_values:
    #             border_value = border_values[m]
    #             m += 1
    #             print("border_value is", border_value)
    # 改成随机参数，减少图片量
    # data_value = random.randrange(1,len(data_values),1)
    # print("data_value is", data_value)
    error_correction_value = random.choice(error_corrections)
    print("ERROR_CORRECT_random_vlue is", error_correction_value)

    box_size_value = random.randrange(2, 3, 1)
    print("box_size_value is", box_size_value)

    border_value = random.randrange(1, 2, 1)
    print("border_value is", border_value)
    data_value = random.choice(data_values)
    print("data_value is", data_value[1:5])
    print("data_value type is", type(data_value))

    qr_code = (version_value + (version_value - 1) * 4 + border_value * 2) * box_size_value
    print("qr_code is", qr_code)
    print("qr_code type is", type(qr_code))
    qr = qrcode.QRCode(version=version_value, error_correction=error_correction_value, box_size=box_size_value,
                       border=border_value, )

    # i = 0
    # for data_value in data_values:
    #     data_value = data_values[i]
    #     print("data_value is", data_value)
    #     qr.add_data(data_value)
    qr.add_data(data_value)
    qr.make(fit=True)
    img = qr.make_image()
    # img.show()
    # time=os.times()
    # print("time is",time)
    # 增加一个将文字写在图片的某个位置
    # 设置所使用的字体
    font = ImageFont.truetype("C:\Windows\Fonts\Arial.ttf", 20)
    # # 打开图片
    # imageFile = "3.jpg"
    # im1 = Image.open(imageFile)

    # 画图
    # draw = ImageDraw.Draw(img)
    # draw.text((0, 0), str(version_value) +'_'+data_value[0:1000]+ '_' + str(
    #     box_size_value) + '_' + str(border_value), False, font=font)  # 设置文字位置/内容/颜色/字体
    # draw = ImageDraw.Draw(img)  # Just draw it!

    # 另存图片
    # img.save("target.jpg")

    img.save(r'C:\Users\zhuoz\Desktop\qr\QR' + '_' + str(qr_code) + '_' + str(version_value) + '_' + data_value[
                                                                                                     0:10] + '_' + str(
        box_size_value) + '_' + str(border_value) + '.jpg')
    # sheet.cell(j*i + 2, 1).value = 'qr'
    # data.save(excelPath)
    # sheet.cell(j*i + 2, 2).value = version_value
    # data.save(excelPath)
    # sheet.cell(j*i + 2, 3).value = data_value
    # data.save(excelPath)
    # sheet.cell(j*i + 2, 4).value = error_correction_value
    # data.save(excelPath)
    # sheet.cell(j*i + 2, 5).value = box_size_value
    # data.save(excelPath)
    # sheet.cell(j*i + 2, 6).value = border_value
    # data.save(excelPath)

    sheet.cell(j + 2, 1).value = 'qr'
    data.save(excelPath)
    sheet.cell(j + 2, 2).value = version_value
    data.save(excelPath)
    sheet.cell(j + 2, 3).value = data_value
    data.save(excelPath)
    sheet.cell(j + 2, 4).value = error_correction_value
    data.save(excelPath)
    sheet.cell(j + 2, 5).value = box_size_value
    data.save(excelPath)
    sheet.cell(j + 2, 6).value = border_value
    data.save(excelPath)
    # i+=1
    j += 1
    # sheet.cell(i+j + 2, 1).value = 'qr'
    # data.save(excelPath)
    # sheet.cell(i+j + 2, 2).value = str(version_value)
    # data.save(excelPath)
    # sheet.cell(i+j + 2, 3).value = data_value
    # data.save(excelPath)
    # sheet.cell(i+j + 2, 4).value = error_correction_value
    # data.save(excelPath)
    # sheet.cell(i+j + 2, 5).value = box_size_value
    # data.save(excelPath)
    # sheet.cell(i+j + 2, 6).value = border_value
    # data.save(excelPath)
''''''
