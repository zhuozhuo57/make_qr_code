import configparser
import os
import re
# import mix as mix
from statistics import mean

import openpyxl

file = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
# print(config)
s = config.read(file + r'\config.ini')
excelPath = config.get("Local", "ResultPath")
filePath = config.get("Local", "logPath")
data = openpyxl.load_workbook(excelPath)
sheet = data['Sheet1']
sheet.cell(1, 1).value = 'date_time'
sheet.cell(1, 2).value = "version"
sheet.cell(1, 3).value = "dataset_path"
count = 0
j = 0
succeed_false_count = 0
succeed_true_count = 0
# image_path_line_count=0
# sheet.cell(1, 4).value = "code_type"
with open(filePath, 'r', encoding="utf-8") as f:
    readlines = f.readlines()
    # print("readlines  is ", readlines)
    for s in readlines:
        # 实现表的项写入
        if ('image_path' in s):
            sheet.cell(4, 1).value = 'image_path'
            data.save(excelPath)
        if ('succeed' in s):
            sheet.cell(4, 2).value = 'succeed'
            data.save(excelPath)
        if ('run_time' in s):
            sheet.cell(4, 3).value = 'run_time'
            data.save(excelPath)
        if ('results' in s):
            sheet.cell(4, 4).value = 'results'
            data.save(excelPath)
        if ('date_time' in s):
            # print("date_time is",s)
            date_time = re.findall(r'"date_time":(.*)"version":(.*),"dataset_path":(.*),', s)
            # print("date_time is" ,date_time)
            # print("date_time[0] is",date_time[0])
            # print("date_time [0][0] is",date_time[0][0])
            # print("version is",date_time[0][1])
            dataset_path = str(date_time[0][2]).strip(',').split(',')[0]
            # print("dataset_path is",dataset_path)
            # 实现提取date_time
            # print(date_time[0])
            # print(type(date_time[0]))
            # 实现date_time_value 写入
            date_time_value = date_time[0][0]
            # print("date_time_value is",date_time_value)
            # print(type((date_time[0])[0]))
            sheet.cell(2, 1).value = date_time_value
            data.save(excelPath)
            # 实现version写入
            version = date_time[0][1]
            # print("version_value is",version)
            # print(type((date_time[0])[1]))
            sheet.cell(2, 2).value = version
            data.save(excelPath)
            # 实现dataset_path写入
            dataset_path = str(date_time[0][2]).strip(',').split(',')[0]
            # print("dataset_path is", dataset_path)
            sheet.cell(2, 3).value = dataset_path
            data.save(excelPath)
            decode_results = re.findall(r'"decode_results":(.*)', s)
            print("decode_results is ", decode_results)
            print("decode_results type is ", type(decode_results))
            # 字符串处理
            decode_result = str(decode_results).strip("['[{").strip("]}']")
            print("decode_result is", decode_result)
            list_decode_result = decode_result.split("},{")
            print("list_decode_result is", list_decode_result)
            for i in range(len(list_decode_result) - 1):
                # print("list_decode_result  len is ", len(list_decode_result))
                print("list_decode_result is", list_decode_result[i])
                print("list_decode_result is", list_decode_result[i].split(","))
                image_name = list_decode_result[i].split(",")[0].split("/")[-1].strip("\"")
                print("image_name is", image_name)
                succeed = list_decode_result[i].split(",")[1].split(":")[1].strip("\"").strip("\"")
                print("succeed is", succeed)
                run_time = list_decode_result[i].split(",")[2].split(":")[1].strip("\"").strip("\"")
                print("run_time is", int(round(float(run_time), 2)))
                print("run_time type is", type(int(round(float(run_time), 2))))
                results = list_decode_result[i].split(",")[3].split(":")[1].strip("\"").strip("\"")
                print("results is", results)
                if (succeed == 'false'):
                    succeed_false_count = i
                    print("succeed_false_count is ", succeed_false_count)
                else:
                    succeed_true_count = i
                    print("succeed_true_count is ", succeed_true_count)
                # 实现数据写入
                sheet.cell(i + 5, 1).value = image_name
                data.save(excelPath)
                sheet.cell(i + 5, 2).value = succeed
                data.save(excelPath)
                sheet.cell(i + 5, 3).value = int(round(float(run_time), 2))
                data.save(excelPath)
                sheet.cell(i + 5, 4).value = results
                data.save(excelPath)
                i += 1
                print("i is", i)
    # # 实现统计行数及结果判断得出解码率
total = ["图片数量（数据集所有图片）", "解码数量（读取成功图片）", "解码正确图片", "解码失败图片", "解码率（解码数量除以图片数量）", "解码正确率（解码正确图除以解码数量）", "最大解码耗时",
         "最小解码耗时", "平均解码耗时", "解码失败类型/数量"]
for j in range(len(total)):
    sheet.cell(j + 5, 5).value = total[j]

    print("total is", total[j])
    # 最大最小平均耗时
    j += 1
    data.save(excelPath)
dd = ["统计", "结果", "单位"]
for k in range(len(dd)):
    sheet.cell(4, k + 5).value = dd[k]
    k += 1
ee = ["张", "张", "张", "张", "%", "%", "ms", "ms", "ms"]
for g in range(len(ee)):
    sheet.cell(g + 5, 7).value = ee[g]
    g += 1


# 从表中读取run_time值 取出最大数。最小数 平均数
def cell_row(xls_path, sheet_name):
    workbook = openpyxl.load_workbook(xls_path)
    worksheek = workbook[sheet_name]
    i = 0
    j = 3
    # 输出所有行的值
    run_time_list = []
    # 计算run_time所有列数值
    for i in range(5, len(list_decode_result) + 4):
        print("run_time_list  len is ", len(list_decode_result) - 1)
        parm = worksheek.cell(row=i, column=j).value
        print(type(parm))
        print(parm)
        run_time_list.append(parm)
    succeed_list = []
    t = 0
    # 计算succeed中false true数值
    for t in range(5, len(list_decode_result) + 4):
        print("succeed  len is ", len(list_decode_result) - 1)
        parms = worksheek.cell(row=t, column=2).value
        print(type(parms))
        print(parms)
        succeed_list.append(parms)
    print("succeed_list", succeed_list)
    # 计算false true 数量
    succeed_false_count = succeed_list.count('false')
    succeed_true_count = succeed_list.count('true')

    # 图片总数
    image_total = len(list_decode_result) - 1
    sheet.cell(5, 6).value = image_total
    # 解码数量
    Number_of_decoding = len(list_decode_result) - 1
    sheet.cell(6, 6).value = Number_of_decoding
    # 解码正确的数量
    image_true_count = succeed_true_count
    sheet.cell(7, 6).value = image_true_count
    # 解码失败的数量
    image_false_count = succeed_false_count
    sheet.cell(8, 6).value = image_false_count
    # 解码率
    decoding_rate = "{0:.0f}".format(Number_of_decoding / image_total * 100)
    sheet.cell(9, 6).value = decoding_rate
    # 解码正确率
    decoding_accuracy = "{0:.0f}".format(image_true_count / Number_of_decoding * 100)
    sheet.cell(10, 6).value = decoding_accuracy
    print("run_time_list is", run_time_list)
    print("run_time_max_col is", len(run_time_list))
    print("run_time_list type is", type(run_time_list))
    # 解码最大耗时
    run_time_max = max(run_time_list)
    print("run_time_max is", run_time_max)
    sheet.cell(11, 6).value = run_time_max
    # 解码最小耗时
    run_time_min = min(run_time_list)
    print("run_time_min is", run_time_min)
    sheet.cell(12, 6).value = run_time_min
    # 解码平均耗时
    run_time_avg = mean(run_time_list)
    print("run_time_avg is", run_time_avg)
    sheet.cell(13, 6).value = run_time_avg
    data.save(excelPath)


cell_row(excelPath, "Sheet1")
