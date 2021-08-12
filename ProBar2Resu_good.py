import configparser
import os
import re

import openpyxl

file = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
print(config)
s = config.read(file + r'\config.ini')
print(s)
excelPath = config.get("Local", "ResultPath")
filePath = config.get("Local", "logPath")
data = openpyxl.load_workbook(excelPath)
sheet = data['Sheet1']
sheet.cell(1, 1).value = 'PicName'
sheet.cell(1, 2).value = "2DDecodeResult"
sheet.cell(1, 3).value = "processTime"
sheet.cell(1, 4).value = "code_type"

file = open(filePath, 'r+', encoding='utf-8')
allLines = file.readlines()
file.seek(0)
count = 0
for line in allLines:
    if ('result : ' in line):
        line = line.replace('result : ', ':null \n')
        print("line is ", line)
        file.write(line)
        # count += 1
        file.close()
    if ('null' in line):
        count += 1
print('count', count)
# print('count is', count)
with open(filePath, 'r', encoding="utf-8") as raw:
    refile = raw.read()
    result = re.findall(r'result :(.*)', refile)
    print(result)
    print(len(result))
    picName = re.findall(r'Image(.*).png', refile)
    ProcssTime = re.findall(r'duration :(.*)', refile)
    code_type = re.findall(r'code_type :(.*)', refile)
for row in range(2, len(picName) + 2):
    sheet.cell(row, 1).value = picName[row - 2]
for row in range(2, len(result) + 2):
    sheet.cell(row, 2).value = result[row - 2]
for row in range(2, len(ProcssTime) + 1):
    sheet.cell(row, 3).value = ProcssTime[row - 2]
for row in range(2, len(code_type) + 1):
    sheet.cell(row, 4).value = code_type[row - 2]
data.save(excelPath)
# add avg

rate = (len(result) - count) / len(picName)
rate1 = (len(result) - count) / len(result)
print(rate1)
if (len(picName) == len(result) & len(ProcssTime) == len(picName)):
    print('the process result is OK, decode rate is $rate', rate)

print(rate)
