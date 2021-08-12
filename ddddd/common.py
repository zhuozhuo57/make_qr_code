#!/usr/bin/python3
# -*-coding:UTF-8 -*-
"""
@File:common.py
@Description:描述
@Author:zhangjian
@Email:zhuozhuo57@163.com
@Date:2020/07/23
"""
import json

import openpyxl


def get_xls_max_nrow(xls_path, sheet_name):
    worksheet = openpyxl.load_workbook(xls_path)
    sheet_name = worksheet[sheet_name]
    max_row = sheet_name.max_row
    max_column = sheet_name.max_column
    print(sheet_name.max_row)
    print(sheet_name.max_column)
    return max_row, max_column
    print(sheet_name)


def get_xls_cell_write_value(xls_path, sheet_name, sheet_name_cell, write_string):
    workbook = openpyxl.load_workbook((xls_path))
    if (sheet_name not in workbook.sheetnames):
        worksheet = workbook.create_sheet(sheet_name, 0)
        worksheet[sheet_name_cell] = write_string
        workbook.save(xls_path)
    else:
        worksheet = workbook[sheet_name]
        worksheet[sheet_name_cell] = write_string
        workbook.save(xls_path)
    s = workbook.sheetnames
    print(s)
    # 判断sheetname是否存在
    if (worksheet.title == "test13"):
        workbook.remove(workbook["test13"])
        workbook.save(xls_path)
    else:
        print("can not find worksheet name ", "test13")


def get_xls_cell_read_value(xls_path, sheet_name, sheet_cell):
    workbook = openpyxl.load_workbook(xls_path)
    worksheet = workbook[sheet_name]
    s = json.loads(worksheet[sheet_cell].value)
    # print(s)
    s_appid = s["appid"]
    # print(s["appid"])
    s_methodname = s["methodname"]
    # print(s["methodname"])
    s_token = ["token"]
    # print(s["token"])
    s_param = s["param"]
    # print(s["param"])
    s_pageNo = json.loads(s_param)["pageNo"]
    # print("s_pageNo is ",s_pageNo)
    # print(json.loads(s_param)["pageNo"])
    s_pageSize = json.loads(s_param)["pageSize"]
    # print(json.loads(s_param)["pageSize"])
    s_operatorId = json.loads(s_param)["operatorId"]
    # print(json.loads(s_param)["operatorId"])
    # return {"s_appid" : s_appid, "s_methodname" : s_methodname,"s_token": s_token,
    #         "s_param" :s_param, "s_pageNo": s_pageNo,"s_pageSize" : s_pageSize,"s_operatorId" : s_operatorId}
    return {"s_appid": s_appid, "s_methodname": s_methodname, "s_token": s_token,
            "s_param": s_param, "s_pageNo": s_pageNo, "s_pageSize": s_pageSize, "s_operatorId": s_operatorId}


def cell(xls_path, sheet_name):
    workbook = openpyxl.load_workbook(xls_path)
    worksheek = workbook[sheet_name]
    # 输出全部值
    for i in range(1, worksheek.max_row + 1):
        for j in range(1, worksheek.max_column + 1):
            parm = worksheek.cell(row=i, column=j).value
            print(type(parm))
            print(parm)


def cell_row(xls_path, sheet_name, j):
    workbook = openpyxl.load_workbook(xls_path)
    worksheek = workbook[sheet_name]
    # 输出所有行的值
    for i in range(1, worksheek.max_row + 1):
        parm = worksheek.cell(row=i, column=j).value
        print(type(parm))
        print(parm)


def cell_colum(xls_path, sheet_name, i):
    workbook = openpyxl.load_workbook(xls_path)
    worksheek = workbook[sheet_name]
    # 输出所有列的值
    for j in range(1, worksheek.max_column + 1):
        parm = worksheek.cell(row=i, column=j).value
        print(type(parm))
        print(parm)


def cell_row_colum(xls_path, sheet_name, i, j):
    workbook = openpyxl.load_workbook(xls_path)
    worksheek = workbook[sheet_name]
    # 输出指定行，指定列列的值
    parm = worksheek.cell(row=i, column=j).value
    print(type(parm))
    print(parm)


methodname = []
tokens = []
appid = []


def cell_colum_to_list(xls_path, sheet_name, j):
    workbook = openpyxl.load_workbook(xls_path)
    worksheek = workbook[sheet_name]

    # 实现取一列某单元格内相同字段名称，输出为列表
    for i in range(2, worksheek.max_row + 1):
        # parm =json.loads(worksheek.cell(row=i, column=j).value)
        parm = worksheek.cell(row=i, column=j).value
        print("parm is", parm)
        print("parm type is", type(parm))
        s_appid = json.loads(parm)["appid"]
        print("s_appid is", s_appid)
        appid.append(s_appid)
        s_methodname = json.loads(parm)["methodname"]
        methodname.append(s_methodname)
        s_token = parm["token"]
        tokens.append(s_token)
        s_pageNo = eval(parm["param"])["pageNo"]
        print(s_pageNo)
        s_pageSize = eval(parm["param"])["pageSize"]
        print(s_pageSize)
        s_operatorId = eval(parm["param"])["operatorId"]
        print(s_operatorId)
        print("methodname is", methodname)
        print("tokens is", tokens)


data01 = {
    "appid": "1",
    "methodname": "queryDeviceList",
    "token": "get_token_value",
    "param": "{\"pageNo\":1,\"pageSize\":10,\"operatorId\":\"1\"}"
}
print("data01 type is", type(data01))
# class getdata():
#     def __init__(self,data):
#         self.data = data
#     def getappid(self):
#         return self.data["appid"]
#     def getmethodname(self):
#         return self.data["methodname"]
#
# class getone(getdata):
#     pass


if __name__ == '__main__':
    # pytest.main([os.path.realpath(__file__), r'--html=F:\tengine2_api_test\Tengine2\result\report.html',
    #              '--self-contained-html'])
    # #get_xls_max_nrow("../data/test1.xlsx","Sheet1")#实现读取表中sheet最大行和最大列的数值
    # #get_xls_cell_write_value("../data/DMcases.xlsx","test","G3","测试")#实现了表内写入内容。删除表的判断
    # #get_xls_cell_read_value("../data/DMcases.xlsx","queryDeviceList","F4")#实现从表中取出各个jason的值
    # tttt()
    # a = get_xls_cell_read_value("../data/DMcases.xlsx","queryDeviceList","F4")#实现从表中取出各个jason的值
    # print(a["s_appid"])
    # a = getdata(data = data01).getmethodname()
    # a = getone(data=data01).getmethodname()
    # print(a)
    # cell("../data/DMcases.xlsx","queryDeviceList")
    # cell_row("../data/DMcases.xlsx","queryDeviceList",4)
    # cell_colum("../data/DMcases.xlsx","queryDeviceList",3)
    # cell_row_colum("../data/DMcases.xlsx","queryDeviceList",2,6)
    cell_colum_to_list("../data/DMcases.xlsx", "queryDeviceList", 6)
