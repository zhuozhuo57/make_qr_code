import configparser
import os
import openpyxl
import qrcode
# '''实现随机生成参数 生成QR码组合字符'''
# version_value = random.randrange(1, 2, 1)
# ERROR_CORRECT_random_vlue = random.randrange(1, 4, 1)
# print("ERROR_CORRECT_random_vlue is", ERROR_CORRECT_random_vlue)
# box_size_value = random.randrange(1, 2, 1)
# print("box_size_value is", box_size_value)
# border_value = random.randrange(99, 100, 1)
# print("version_value is", version_value)
# qr = qrcode.QRCode(version=21, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=box_size_value,
#                    border=border_value, )
'''实现xlsx 文件路径'''
file = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
print(config)
s = config.read(file + r'\config.ini')
print(s)
excelPath = config.get("QRLocal", "QRResultPath")
filePath = config.get("QRLocal", "QRlogPath")
data = openpyxl.load_workbook(excelPath)
sheet = data['Sheet1']
''' #码制	Version版本号	内容	密度	长度	密度mm	长度mm	浓度'''
sheet.cell(1, 1).value = '码制qr'
sheet.cell(1, 2).value = "Version版本号"
sheet.cell(1, 3).value = "内容add_data"
sheet.cell(1, 4).value = "容错率ERROR_CORRECT"
sheet.cell(1, 5).value = "像素尺寸box_size"  # box_size表示每个模块的像素数，“每边模块数” × box_size 就是二维码的像素尺寸了。公式如下：
# (21 + (version - 1) * 4 + border * 2) * box_size
sheet.cell(1, 6).value = "每边模块数量border_value"  # 边距是以模块数量为单位，比如border=4就是表示4个模块。
# sheet.cell(1, 7).value = "浓度Concentration"#浓度  70% 50% 30%.
data.save(excelPath)
# pixel_size
''''''
# '''实现list中读取内容生成QR码保存图片'''
# data = ['1111111', 'abcder', 'ABCDabcd','!@#$%^&*(){}_+|~:"<>?']
data = ['1111111', 'abcder', 'ABCDabcd', '!@#$%^&(){}_+~']
version_values = [20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40]
# error_corrections=["ERROR_CORRECT_L","ERROR_CORRECT_M","ERROR_CORRECT_Q","ERROR_CORRECT_H"]
error_corrections = [1, 0, 3, 2]
box_size_values = [10, 20, 30, 40, 50]
border_values = [1, 3, 5, 7, 9]
i = 0
for data_value in data:
    data_value = data[i]
    print("data_value is", data_value)
    i += 1
    j = 0
    for version_value in version_values:
        version_value = version_values[j]
        j += 1
        print("version_value is", version_value)
        k = 0
        for error_correction_value in error_corrections:
            error_correction_value = error_corrections[k]
            k += 1
            print("error_correction is", error_correction_value)
            l = 0
            for box_size_value in box_size_values:
                box_size_value = box_size_values[l]
                l += 1
                print("box_size_value is", box_size_value)
                m = 0
                for border_value in border_values:
                    border_value = border_values[m]
                    m += 1
                    print("border_value is", border_value)
                    qr = qrcode.QRCode(version=version_value, error_correction=error_correction_value, box_size=box_size_value,
                                       border=border_value, )
                    qr.add_data(data_value)
                    qr.make(fit=True)
                    img = qr.make_image()
                    # img.show()
                    # time=os.times()
                    # print("time is",time)
                    img.save(r'C:\Users\zhuoz\Desktop\qr\QR' + '_' + str(version_value) + '_' + data_value + '_' + str(
                        box_size_value) + '_' + str(border_value) + '.jpg')

''''''
