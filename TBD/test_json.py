import configparser
import json
import os
from statistics import mean

import openpyxl

file = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
s = config.read(file + r'\config.ini')
excelPath = config.get("Local", "ResultPath")
filePath = config.get("Local", "logPath")
data = openpyxl.load_workbook(excelPath)
sheet = data['Sheet1']
#定义写入row name 集合
test_data=["image_path","image_name","succeed","run_time","type","result","target_result","rect_points","target_rect_points"]
for i in range(len(test_data)):
    sheet.cell(1, i+1).value = test_data[i]
    i+=1
    # for s in readlines:
    #     # 实现表的项写入
    #     if ('image_path' in s):
    #         sheet.cell(4, 1).value = 'image_path'
    #         #data.save(excelPath)
    #     if ('succeed' in s):
    #         sheet.cell(4, 2).value = 'succeed'
    #         #data.save(excelPath)
    #     if ('run_time' in s):
    #         sheet.cell(4, 3).value = 'run_time'
    #         #data.save(excelPath)
    #     if ('results' in s):
    #         sheet.cell(4, 4).value = 'results'
    #        # data.save(excelPath)
    #     if ('type' in s):
    #         sheet.cell(4, 5).value = 'type'
    #     if ('result' in s):
    #         sheet.cell(4, 6).value = 'result'
    #         #data.save(excelPath)
    #     if ('rect_points' in s):
    #         sheet.cell(4, 7).value = 'rect_points'
    #         #data.save(excelPath)
    #     if ('date_time' in s):
    #         # print("date_time is",s)
    #         date_time = re.findall(r'"date_time":(.*)"version":(.*),"dataset_path":(.*),', s)
    #         # print("date_time is" ,date_time)
    #         # print("date_time[0] is",date_time[0])
    #         # print("date_time [0][0] is",date_time[0][0])
    #         # print("version is",date_time[0][1])
    #         dataset_path = str(date_time[0][2]).strip(',').split(',')[0]
    #         # print("dataset_path is",dataset_path)
    #         # 实现提取date_time
    #         # print(date_time[0])
    #         # print(type(date_time[0]))
    #         # 实现date_time_value 写入
    #         date_time_value = date_time[0][0]
    #         # print("date_time_value is",date_time_value)
    #         # print(type((date_time[0])[0]))
    #         sheet.cell(2, 1).value = date_time_value
    #         data.save(excelPath)
    #         # 实现version写入
    #         version = date_time[0][1]
    #         # print("version_value is",version)
    #         # print(type((date_time[0])[1]))
    #         sheet.cell(2, 2).value = version
    #         data.save(excelPath)
    #         # 实现dataset_path写入
    #         dataset_path = str(date_time[0][2]).strip(',').split(',')[0]
    #         # print("dataset_path is", dataset_path)
    #         sheet.cell(2, 3).value = dataset_path
    #        # data.save(excelPath)
    #         decode_results = re.findall(r'"decode_results":(.*)', s)
    #         #print("decode_results is ", decode_results)
    #         #print("decode_results type is ", type(decode_results))
    #         # 字符串处理
    #         decode_result = str(decode_results).strip("['[{").strip("]}']")
    #         #print("decode_result is", decode_result)
    #         list_decode_result = decode_result.split(",")
    #         #print("list_decode_result is", list_decode_result)
    #         # for i in range(len(list_decode_result)):
    #         #     # print("list_decode_result  len is ", len(list_decode_result))
    #         #    # print("list_decode_result is", list_decode_result[i])
    #         #     #print("list_decode_result is", list_decode_result[i].split(","))
    #         #     image_name = list_decode_result[i].split(",")[0].split("/")[-1].strip("\"")
    #         #     #print("image_name is", image_name)
    #         #     succeed = list_decode_result[i].split(",")[1].split(":")[1].strip("\"").strip("\"")
    #         #     #print("succeed is", succeed)
    #         #     run_time = list_decode_result[i].split(",")[2].split(":")[1].strip("\"").strip("\"")
    #         #     #print("run_time is", int(round(float(run_time), 2)))
    #         #     #print("run_time type is", type(int(round(float(run_time), 2))))
    #         #     #'"results":[
    #         #     # {"type":"qr"', '"result":"DM123456781234567890"', '"rect_points":
    #         #     # [
    #         #     # [825.4306640625', '586.499755859375]', '[938.5455322265625', '577.678466796875]', '[947.5989990234375', '693.7711181640625]', '[834.484130859375', '702.5924072265625]
    #         #     # ]
    #         #     # }
    #         #     # ]'
    #         #     print("list_decode_result [0] is", list_decode_result[i].split(',')[0])
    #         #     print("list_decode_result [1] is", list_decode_result[i].split(',')[1])
    #         #     print("list_decode_result [2] is", list_decode_result[i].split(',')[2])
    #         #     if(list_decode_result[i].split(',')[3]=='"results":[]'):
    #         #         continue
    #         #     print("list_decode_result [3][0] is", list_decode_result[i].split(',')[3].split(":")[0].strip("\"").strip("\""))
    #         #     print("list_decode_result [3][1] is", list_decode_result[i].split(',')[3].split(":")[1].strip("\[").strip("\{").strip("\"").strip("\""))
    #         #     print("list_decode_result [3][2] is", list_decode_result[i].split(',')[3].split(":")[2].strip("\"").strip("\""))
    #         #     print("list_decode_result [4] is", list_decode_result[i].split(',')[4].split(":")[0].strip("\"").strip("\""))
    #         #     print("list_decode_result [4] is", list_decode_result[i].split(',')[4].split(":")[1].strip("\"").strip("\""))
    #         #     #if(list_decode_result[i].split(',')[5].split(":")[0].strip("\"").strip("\"")=='"results":[]')
    #         #     print("list_decode_result [5] is", list_decode_result[i].split(',')[5])
    #         #     print("list_decode_result [5] is", list_decode_result[i].split(',')[5].split(":")[0].strip("\"").strip("\""))
    #         #     print("list_decode_result [5] is", list_decode_result[i].split(',')[5].split(":")[1].strip("\[").strip("\["))
    #         #     print("list_decode_result [6] is", list_decode_result[i].split(',')[6].strip("\]"))
    #         #     print("list_decode_result [7] is", list_decode_result[i].split(',')[7].strip("\["))
    #         #     print("list_decode_result [8] is", list_decode_result[i].split(',')[8].strip("\]"))
    #         #     print("list_decode_result [9] is", list_decode_result[i].split(',')[9].strip("\["))
    #         #     print("list_decode_result [10] is", list_decode_result[i].split(',')[10].strip("\]"))
    #         #     print("list_decode_result [11] is", list_decode_result[i].split(',')[11].strip("\["))
    #         #     print("list_decode_result [12] is", list_decode_result[i].split(',')[12].strip("\]").strip("\]").strip("\}").strip("\]"))
    #         #
    #         #
    #         #
    #         #     #type= list_decode_result[i].split(",")[3].split(":")[0]
    #         #     #type1= list_decode_result[i].split(",")[3][1]
    #         #     # print("type is", type)
    #         #     # #print("type is", type1)
    #         #     results =  list_decode_result[i].split(",")[3].split(":")[1].strip("\"").strip("\"")
    #         #     # print("results is", results)
    #         #
    #         #
    #         #
    #         #
    #         #
    #         #
    #         #
    #         #
    #         #
    #         #
    #         #
    #         #
    #         #     if (succeed == 'false'):
    #         #         succeed_false_count = i
    #         #         print("succeed_false_count is ", succeed_false_count)
    #         #     else:
    #         #         succeed_true_count = i
    #         #         print("succeed_true_count is ", succeed_true_count)
    #         #     # 实现数据写入
    #         #     sheet.cell(i + 5, 1).value = image_name
    #         #     #data.save(excelPath)
    #         #     sheet.cell(i + 5, 2).value = succeed
    #         #    # data.save(excelPath)
    #         #     sheet.cell(i + 5, 3).value = int(round(float(run_time), 2))
    #         #     #data.save(excelPath)
    #         #     sheet.cell(i + 5, 4).value = results
    #         #     #data.save(excelPath)
    #         #     i += 1
    #         #     #print("i is", i)
    # # 实现统计行数及结果判断得出解码率
reslut_title = ["测试时间","算法版本","测试图片集","解码率","正确率","耗时","最大耗时","最小耗时","扫码失败总数","失败文件路径分类"]
for j in range(len(reslut_title)):
    sheet.cell(1, j+len(test_data)).value = reslut_title[j]
    j += 1


#定义单位等
# dd = ["统计", "结果", "单位"]
# for k in range(len(dd)):
#     sheet.cell(4, k + 5).value = dd[k]
#     k += 1
# ee = ["张", "张", "张", "张", "%", "%", "ms", "ms", "ms"]
# for g in range(len(ee)):
#     sheet.cell(g + 5, 7).value = ee[g]
#     g += 1
#处理测试日志文件，提取json数据
with open(filePath, 'r', encoding="utf-8") as f:
    readlines = f.read()
    s = json.loads(readlines)
    # 实现数据写入
    print("date_time  is",s["date_time"])
    sheet.cell(2, 9).value = s["date_time"]
    print("version is ",s["version"])
    sheet.cell(2, 10).value = s["version"]
    print("dataset_path is",s["dataset_path"])
    sheet.cell(2, 11).value = s["dataset_path"]
    father_dir = []
    child_dir = []
    list_fail_image_path=[]
    for i in range(len(s["decode_results"])):
        # print("decode_results "%s,i,"is",s["decode_results"][i])
        print("image_path is",str(s["decode_results"][i]["image_path"]))
        sheet.cell(i+2, 1).value = str(s["decode_results"][i]["image_path"])
        print("image_name is",str(s["decode_results"][i]["image_path"]).split("/")[-1])
        sheet.cell(i + 2, 2).value = str(s["decode_results"][i]["image_path"]).split("/")[-1]
        print("succeed is",s["decode_results"][i]["succeed"])
        sheet.cell(i+2, 3).value = s["decode_results"][i]["succeed"]
        if (str(s["decode_results"][i]["succeed"]) == 'false'):
            #返回解码失败路径
            fail_image_path=str(s["decode_results"][i]["image_path"])
            list_fail_image_path.append(fail_image_path)
            print("根据succed返回值获取 图片路径", str(s["decode_results"][i]["image_path"]))
            #解析截取路径名称
            print("根据succed返回值获取 获取父目录", str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[0:2])
            print("根据succed返回值获取 获取父目录", str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[0]+str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[1])
            # print("根据succed返回值获取 获取父目录", str(s["decode_results"][i]["image_path"]).split("/")[2:4])
            print("根据succed返回值获取 获取子目录", str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[1])
            # 图片父目录

            father_dir.append((str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[0]+"/"+str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[1]))
            print("father_dir is", father_dir)
            # 图片子目录

            child_dir.append(str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[1])
            print("child_dir is", child_dir)
            #print("根据succed返回值获取 图片路径", str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[2])
        print("run_time is",s["decode_results"][i]["run_time"])
        sheet.cell(i+2, 4).value = int(s["decode_results"][i]["run_time"])
        #print("results is",s["decode_results"][i]["results"])
        for j in range(len(s["decode_results"][i]["results"])):
            #print("decode_results "%s,i,"is",s["decode_results"][i]["results"][j])
            print("type is",s["decode_results"][i]["results"][j]["type"])
            sheet.cell(i + 2, 5).value = s["decode_results"][i]["results"][j]["type"]
            print("result is",s["decode_results"][i]["results"][j]["result"])
            sheet.cell(i + 2, 6).value = s["decode_results"][i]["results"][j]["result"]
            print("rect_points is",s["decode_results"][i]["results"][j]["rect_points"])
            sheet.cell(i + 2, 8).value = format(s["decode_results"][i]["results"][j]["rect_points"])
            j+=1
        i+=1
data.save(excelPath)
#计算解码率，正确率，平均耗时，最大耗时，最小耗时，失败类型数量分布
#解码率
#计算方法：len(succeed=ture)/len(s["decode_results"])
# 从表中读取run_time值 取出最大数。最小数 平均数
def cell_row(xls_path, sheet_name):
    workbook = openpyxl.load_workbook(xls_path)
    worksheek = workbook[sheet_name]
    # 输出所有行的值
    run_time_list = []
    # 计算run_time所有列数值
    for i in range(2, len(s["decode_results"])):
        #print("run_time_list  len is ", len(list_decode_result) - 1)
        parm = worksheek.cell(row=i, column=4).value
#        print(type(parm))
        #print(parm)
        run_time_list.append(parm)
    succeed_list = []
    # 计算succeed中false true数值
    for i in range(2, len(s["decode_results"])):
        #print("succeed  len is ", len(list_decode_result) - 1)
        parms = worksheek.cell(row=i, column=3).value
        #print(type(parms))
        #print(parms)
        succeed_list.append(parms)
    # 计算false true 数量
    succeed_false_count = succeed_list.count('false')
    sheet.cell(2, len(test_data)+8).value = succeed_false_count
    succeed_true_count = succeed_list.count('true')
    # # 图片总数
    image_total = len(s["decode_results"])
    # 解码数量
    Number_of_decoding = len(s["decode_results"])
    # 解码率
    decoding_rate = "{:.2%}".format(Number_of_decoding / image_total)
    sheet.cell(2, len(test_data)+3).value = decoding_rate
    # 解码正确率
    decoding_accuracy = "{:.2%}".format(succeed_true_count / Number_of_decoding)
    sheet.cell(2, len(test_data)+4).value = decoding_accuracy
    # 解码最大耗时
    run_time_max = max(run_time_list)
    sheet.cell(2,len(test_data)+6).value = run_time_max
    # 解码最小耗时
    run_time_min = min(run_time_list)
    sheet.cell(2,len(test_data)+7).value = run_time_min
    # 解码平均耗时
    run_time_avg = mean(run_time_list)
    sheet.cell(2,len(test_data)+5).value =run_time_avg


cell_row(excelPath, "Sheet1")

print("list_fail_image_path is",list_fail_image_path)
print("father_dir is",father_dir)
set(father_dir)
father_dir.sort()
print("sort  is",father_dir)
print("set(father_dir) is",set(father_dir))
sdfasd_list = []
for i in range(len(set(father_dir))):

    # print("set(father_dir) i is",list(list(set(father_dir))[i]))
    # print("set(father_dir) i is",list(list(set(father_dir)).sort()[i]))
    #print("set(father_dir))is",)
    list(set(father_dir)).sort()
    print("set(father_dir))is",list(set(father_dir)))
    print("set(father_dir))is",list(set(father_dir))[i])

    #计算count
    father_dir.count(list(set(father_dir))[i])
    print("list(set(father_dir))[i]  count is",father_dir.count(list(set(father_dir))[i]))


    #合并写入
    print("和并写入单行",(list(set(father_dir))[i]+":"+str(father_dir.count(list(set(father_dir))[i]))))
    m=list(set(father_dir))[i]+":"+str(father_dir.count(list(set(father_dir))[i]))


    sdfasd_list.append(m)
    #str_sss=str(sdfasd_list)
    print("sdfasd_list  is ",str(sdfasd_list))
    print("sdfasd_list  is ",str(sdfasd_list).strip("[").strip("]").strip("'"))
    print("sdfasd_list  is ",str(sdfasd_list).strip("[").strip("]").strip("'").replace("', '", "\r\n"))



    i+=1
sheet.cell(2, len(test_data) + 9).value = str(sdfasd_list).strip("[").strip("]").strip("'").replace("', '", "\r\n")
data.save(excelPath)


# strsss =[]
# for i in range(len(sdfasd_list)):
#     ss = sdfasd_list[i] + "\r\n"
#     print("+++++++++++++++++  is ", sdfasd_list[i]+"\r\n")
#
#
#     strsss.append(ss)
#
#     i+=1
#
# print("_____________ is" ,str(strsss).replace(",","\r\n"))
# sheet.cell( 2, len(test_data) + 9).value = str(strsss).replace(",","\r\n")

print("类型",father_dir)
print("数量",len(father_dir))
#定义拼接字符串list

#print("sdfasd_list  is ", str(str(sdfasd_list).split(",")))
