import configparser
import os

import openpyxl

file = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
s = config.read(file + r'\config.ini')
excelPath = config.get("Local", "ResultPath")
filePath = config.get("Local", "logPath")
data = openpyxl.load_workbook(excelPath)
sheet = data['Sheet1']
sheet.cell(1, 1).value = 'imageid'
sheet.cell(1, 2).value = "main_dir"
sheet.cell(1, 3).value = "scend_dir"
sheet.cell(1, 4).value = "picname"
sheet.cell(1, 5).value = "2DDecodeResult"
sheet.cell(1, 6).value = "duration"
sheet.cell(1, 7).value = "code_type"

# file = open(filePath, 'r+', encoding='utf-8')
# allLines = file.readlines()
# file.seek(0)
# for line in allLines:
#     if (': \n' in line):
#         line = line.replace(': ', ':null \n')
#         print("line is ", line)
#     file.write(line)
# file.close()
i = 0
with open(filePath, 'r', encoding="utf-8") as raw:
    refile = raw.readlines()
    print("refile is ", type(refile))
    # del_=str.replace(r'-------------------------------------------', '')
    # print(del_)
    # add_null=str.replace(r'result :','result :null')
    # print(add_null)
    # for row in range(2, len(result) + 1):
    #     sheet.cell(row, 5).value = result[row - 2]
    # ddd = re.findall(r'Image(.*)', refile)
    # image = re.findall(r'Image\[(.*)\] : ', raw)
    # print("image is", image)
    # for row in range(2, len(image) + 1):
    #     sheet.cell(row, 1).value = image[row - 2]
    # #print("image is",image)
    # dir = re.findall(r'/mnt/DM//(.*)(.*)/', refile)
    # #print("dir",dir)
    # dir2 = re.findall(r'/mnt/DM//(.?.?.?)/(.*)/', refile)
    # print("dir2 is",dir2)
    # png = re.findall(r'/(.*?)/(.*?)//(.*?)/(.*?)/(.*)', raw)
    # print("png  is",png)
#
# for str in png:
#     dir=png[len(png)-1][2]
#     print("dir is",dir)
#     for row in range(2, len(dir) + 1):
#         sheet.cell(row, 2).value = dir
#     dir_type=png[len(png)-1][3]
#     print("dir_type is",dir_type)
#     for row in range(2, len(dir_type) + 1):
#         sheet.cell(row, 3).value = dir_type
#     png_name=png[len(png)-1][4]
#     print(png[len(png)-1][4])
#     for row in range(2, len(png_name) + 1):
#         sheet.cell(row, 4).value = png_name
#     duration = re.findall(r'duration :(.*)ms', refile)
#     code_type = re.findall(r'code_type :(.*)', refile)
#
#     for row in range(2, len(duration) + 1):
#         sheet.cell(row, 6).value = duration[row - 2]
#     for row in range(2, len(code_type) + 1):
#         sheet.cell(row, 7).value = code_type[row - 2]
# data.save(excelPath)


# data.save(excelPath)
# print("picName type is", type(ddd))
# str = str(ddd).split(",")
# i = 0
# j = 0
# kk = ''
# m = ''
# b = ''
# while kk in str[i]:
#     #     print("str is", i, str[i])
#
#     b = str[i].split(':')
#     print("b is", b)
#     print("image id number is", b[0][3:len(b[0]) - 1].strip(']'))
#     print("b[1] is", b[1])
#     c = b[1].split("/")
#     print("c is", c)
#     while m in c[len(c)-1]:
#         imageid = b[0][3:len(b[0]) - 1].strip(']')
#         print('imageid is', imageid)
#         main_dir = c[4]
#         print("dir is", c[4])
#         scend_dir = c[5]
#         print("dir2 is", c[5])
#         picname = b[0][3:len(b[0]) - 1].strip(']')
#         print("png-name is", c[len(c) - 1].strip('\']'))
#         if (j == len(c) - 1):
#             break
#         j += 1
#     i += 1
#     if (i == len(str)):
#         break
#     if (i == 0):
#         print("str[0] is", str[0].replace('[\'', '\''))
#     if (i == len(str)):
#         print("str[len(str)] is", str[len(str)].replace('\']', '\''))


# for row in range(2, len(imageid) + 2):
#     sheet.cell(row, 1).value = imageid[row - 2]
# for row in range(2, len(main_dir) + 2):
#     sheet.cell(row, 2).value = main_dir[row - 2]
# for row in range(2, len(scend_dir) + 2):
#     sheet.cell(row, 3).value = scend_dir[row - 2]
# for row in range(2, len(picname) + 2):
#     sheet.cell(row, 4).value = picname[row - 2]
# for row in range(2, len(result) + 2):
#     sheet.cell(row, 5).value = result[row - 2]
