# -*- coding=utf-8 -*-
import configparser
import json
from statistics import mean
import time, os
import openpyxl


file = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
s = config.read(file + r'\config.ini')
excelPath = config.get("Local", "ResultPath")
filePath = config.get("Local", "logPath")
# data = openpyxl.load_workbook(excelPath)
# sheet = data['Sheet1']

# 处理测试日志文件，提取json数据
with open(filePath, 'r', encoding="utf-8") as f:
    readlines = f.read()
    s = json.loads(readlines)
    # 实现数据写入
    print("date_time  is", s["date_time"])
    # sheet.cell(2, 9).value = s["date_time"]
    verison = s["version"]
    print("version is ", s["version"])

    # sheet.cell(2, 10).value = s["version"]
    print("dataset_path is", s["dataset_path"])
    # sheet.cell(2, 11).value = s["dataset_path"]
    father_dir = []
    child_dir = []
    list_fail_image_path = []
    for i in range(len(s["decode_results"])):
        # print("decode_results "%s,i,"is",s["decode_results"][i])
        print("image_path is", str(s["decode_results"][i]["image_path"]))
        # sheet.cell(i+2, 1).value = str(s["decode_results"][i]["image_path"])
        print("image_name is", str(s["decode_results"][i]["image_path"]).split("/")[-1])
        # sheet.cell(i + 2, 2).value = str(s["decode_results"][i]["image_path"]).split("/")[-1]
        print("succeed is", s["decode_results"][i]["succeed"])
        # sheet.cell(i+2, 3).value = s["decode_results"][i]["succeed"]
        if (str(s["decode_results"][i]["succeed"]) == 'false'):
            # 返回解码失败路径
            fail_image_path = str(s["decode_results"][i]["image_path"])
            list_fail_image_path.append(fail_image_path)
            print("根据succed返回值获取 图片路径", str(s["decode_results"][i]["image_path"]))
            # 解析截取路径名称
            print("根据succed返回值获取 获取父目录",
                  str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[0:2])
            print("根据succed返回值获取 获取父目录",
                  str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[0] +
                  str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[1])
            # print("根据succed返回值获取 获取父目录", str(s["decode_results"][i]["image_path"]).split("/")[2:4])
            print("根据succed返回值获取 获取子目录",
                  str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[1])
            # 图片父目录
            father_dir.append((str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[
                                   0] + "/" +
                               str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[1]))
            print("father_dir is", father_dir)
            # 图片子目录
            child_dir.append(str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[1])
            print("child_dir is", child_dir)
            # print("根据succed返回值获取 图片路径", str(s["decode_results"][i]["image_path"]).strip(str(s["dataset_path"])).split("/")[2])
        print("run_time is", s["decode_results"][i]["run_time"])
        # sheet.cell(i+2, 4).value = int(s["decode_results"][i]["run_time"])
        # print("results is",s["decode_results"][i]["results"])
        for j in range(len(s["decode_results"][i]["results"])):
            # print("decode_results "%s,i,"is",s["decode_results"][i]["results"][j])
            print("type is", s["decode_results"][i]["results"][j]["type"])
            sheetname_type = s["decode_results"][i]["results"][j]["type"]
            print("sheetname_type   is", sheetname_type)
            print("sheetname_type  type is", type(sheetname_type))
            # sheet.cell(i + 2, 5).value = s["decode_results"][i]["results"][j]["type"]
            print("result is", s["decode_results"][i]["results"][j]["result"])
            # sheet.cell(i + 2, 6).value = s["decode_results"][i]["results"][j]["result"]
            print("rect_points is", s["decode_results"][i]["results"][j]["rect_points"])
            # sheet.cell(i + 2, 8).value = format(s["decode_results"][i]["results"][j]["rect_points"])
            j += 1
        i += 1
# 计算解码率，正确率，平均耗时，最大耗时，最小耗时，失败类型数量分布
# 解码率
# 计算方法：len(succeed=ture)/len(s["decode_results"])
# 从表中读取run_time值 取出最大数。最小数 平均数
def cell_row(xls_path, sheet_name):
    workbook = openpyxl.load_workbook(xls_path)
    workbook.create_sheet(sheet_name)
    worksheek = workbook[sheet_name]
    with open(filePath, 'r', encoding="utf-8") as f:
        readlines = f.read()
        s = json.loads(readlines)
    # 输出所有行的值
    run_time_list = []
    # 计算run_time所有列数值
    for i in range(2, len(s["decode_results"])):
        # print("run_time_list  len is ", len(list_decode_result) - 1)
        parm = worksheek.cell(row=i, column=4).value
        #        print(type(parm))
        # print(parm)
        run_time_list.append(parm)
    succeed_list = []
    # 计算succeed中false true数值
    for i in range(2, len(s["decode_results"])):
        parms = worksheek.cell(row=i, column=3).value
        succeed_list.append(parms)
    # 计算false true 数量
    succeed_false_count = succeed_list.count('false')
    # sheet.cell(2, len(test_data)+8).value = succeed_false_count
    succeed_true_count = succeed_list.count('true')
    # # 图片总数
    image_total = len(s["decode_results"])
    # 解码数量
    Number_of_decoding = len(s["decode_results"])
    # 解码率
    decoding_rate = "{:.2%}".format(Number_of_decoding / image_total)
    # sheet.cell(2, len(test_data)+3).value = decoding_rate
    # 解码正确率
    decoding_accuracy = "{:.2%}".format(succeed_true_count / Number_of_decoding)
    # sheet.cell(2, len(test_data)+4).value = decoding_accuracy
    # 解码最大耗时
    run_time_max = max(run_time_list)
    # sheet.cell(2,len(test_data)+6).value = run_time_max
    # 解码最小耗时
    run_time_min = min(run_time_list)
    # sheet.cell(2,len(test_data)+7).value = run_time_min
    # 解码平均耗时
    run_time_avg = mean(run_time_list)
    # sheet.cell(2,len(test_data)+5).value =run_time_avg
cell_row(excelPath, "Sheet1")
print("list_fail_image_path is", list_fail_image_path)
print("father_dir is", father_dir)
set(father_dir)
father_dir.sort()
print("sort  is", father_dir)
print("set(father_dir) is", set(father_dir))
sdfasd_list = []
for i in range(len(set(father_dir))):
    list(set(father_dir)).sort()
    print("set(father_dir))is", list(set(father_dir)))
    print("set(father_dir))is", list(set(father_dir))[i])
    # 计算count
    father_dir.count(list(set(father_dir))[i])
    print("list(set(father_dir))[i]  count is", father_dir.count(list(set(father_dir))[i]))
    # 合并写入
    print("和并写入单行", (list(set(father_dir))[i] + ":" + str(father_dir.count(list(set(father_dir))[i]))))
    m = list(set(father_dir))[i] + ":" + str(father_dir.count(list(set(father_dir))[i]))
    sdfasd_list.append(m)
    print("sdfasd_list  is ", str(sdfasd_list))
    print("sdfasd_list  is ", str(sdfasd_list).strip("[").strip("]").strip("'"))
    print("sdfasd_list  is ", str(sdfasd_list).strip("[").strip("]").strip("'").replace("', '", "\r\n"))
    i += 1
# sheet.cell(2, len(test_data) + 9).value = str(sdfasd_list).strip("[").strip("]").strip("'").replace("', '", "\r\n")
# data.save(excelPath)

print("类型", father_dir)
print("数量", len(father_dir))
# 定义拼接字符串list
# print("sdfasd_list  is ", str(str(sdfasd_list).split(",")))
# 抽取变量
image_total = len(s["decode_results"])
Number_of_decoding = len(s["decode_results"])
decoding_rate = "{:.2%}".format(Number_of_decoding / image_total)
succeed_list = []
succeed_true_count = succeed_list.count('true')
decoding_accuracy = "{:.2%}".format(succeed_true_count / Number_of_decoding)
# run_time_list = ['']
# run_time_avg = mean(run_time_list)
# run_time_max = max(run_time_list)
# run_time_min = min(run_time_list)
succeed_false_count = succeed_list.count('false')
fail_path_total = str(sdfasd_list).strip("[").strip("]").strip("'").replace("', '", "\r\n")
# 实现html 显示

class Template_mixin(object):
    """html报告"""
    HTML_TMPL ="""
<!DOCTYPE html>
<html lang="en">
<body>
<head>
<meta charset="UTF-8">
<title>算法自动化测试报告</title>
<link href="http://libs.baidu.com/bootstrap/3.0.3/css/bootstrap.min.css" rel="stylesheet">
<h2 style="font-family: Microsoft YaHei">算法自动化测试报告</h2>
<p class='attribute'><strong>测试结果 : </strong> %(value)s</p>
<style type="text/css" media="screen">
body  { font-family: Microsoft YaHei,Tahoma,arial,helvetica,sans-serif;padding: 20px;}
</style>
</head>
<table id='result_table' class="table table-condensed table-bordered table-hover">
    <colgroup>
       <col align='left' />
       <col align='right' />
       <col align='right' />
       <col align='right' />
    </colgroup>
    <tr id='header_row' class="text-center success" style="font-weight: bold;font-size: 14px;">
        <th width="50px">commit_times_id</th>
        <th width="50px">测试时间</th>
        <th width="50px">提交代码commit_id</th>
        <th width="50px">测试数据集</th>
        <th width="50px">测试结果</th>
    </tr>
    %(table_tr2)s
</table>
<table id='result_table' class="table table-condensed table-bordered table-hover">
    <tr id='header_row' class="text-center success" style="font-weight: bold;font-size: 14px;">
    <colgroup>
       <col align='left' />
       <col align='right' />
       <col align='right' />
       <col align='right' />
       <col align='right' />
    </colgroup>
    <tr id='header_row' class="text-center success" style="font-weight: bold;font-size: 14px;">
        <th width="50px">reslut_set_id</th>
        <th width="50px">测试时间</th>
        <th width="50px">算法版本</th>
        <th width="50px">测试图片集</th>
        <th width="50px">解码率</th>
        <th width="50px">正确率</th>
        <th width="50px">平均耗时</th>
        <th width="50px">最大耗时</th>
        <th width="50px">最小耗时</th>
        <th width="50px">扫码失败总数</th>
        <th width="50px">扫码失败文件分布</th>
    </tr>
    %(table_tr)s
</table>
<details>
<summary>点击查看详细测试结果</summary>
<table id='result_table' class="table table-condensed table-bordered table-hover">
    <tr id='header_row' class="text-center success" style="font-weight: bold;font-size: 14px;">
        <th width="50px">id</th>
        <th width="50px">图片路径</th>
        <th width="50px">解码是否成功</th>
        <th width="50px">图片处理耗时</th>
        <th width="50px">图片解码类型</th>
        <th width="50px">图片解码内容</th>
        <th width="50px">图片原本标注信息</th>
        <th width="50px">检测坐标</th>
    </tr>
    %(table_tr3)s
</table>
</details>
</body>
</html>"""
    # 总数据
    TABLE_TMPL_TOTAL = """
        <tr class='failClass warning'>
            <td width="50px">%(commit_times_id)s</td>
            <td width="50px">%(测试时间)s</td>
            <td width="50px">%(提交代码commit_id)s</td>
            <td width="50px">%(测试数据集)s</td>
            <td width="50px">%(测试结果)s</td>
        </tr>"""
    # 详情表头
    TABLE_TMPL_MODULE = """
        <tr id='header_row' class="failClass warning" style="font-weight: bold;font-size: 14px;">
            <td width="50px">%(reslut_set_id)s</td>
            <td width="50px">%(测试时间)s</td>
            <td width="50px">%(算法版本)s</td>
            <td width="50px">%(测试图片集)s</td>
            <td width="50px">%(解码率)s</td>
            <td width="50px">%(正确率)s</td>
            <td width="50px">%(耗时)s</td>
            <td width="50px">%(最大耗时)s</td>
            <td width="50px">%(最小耗时)s</td>
            <td width="50px">%(扫码失败总数)s</td>
            <td width="50px">%(扫码失败文件分布)s</td>
        </tr>"""
    # case数据
    TABLE_TMPL_CASE = """
        <tr id='header_row' class="failClass warning" style="font-weight: bold;font-size: 14px;">
            <td width="50px">%(id)s</td>
            <td width="50px">%(image_path)s</td>
            <td width="50px">%(succeed)s</td>
            <td width="50px">%(run_time)s</td>
            <td width="50px">%(type)s</td>
            <td width="50px">%(result)s</td>
            <td width="50px">%(target_result)s</td>
            <td width="50px">%(rect_points)s</td>
        </tr>"""


if __name__ == '__main__':
    table_tr0 = ''
    table_tr1 = ""
    table_tr2 = ""
    numfail = 1
    numsucc = 9
    html = Template_mixin()
    commit_times_id = 1
    test_result_data = {"commit_times_id": commit_times_id,
                        "测试时间": '{date}'.format(date=time.strftime('%Y%m%d%H%M%S')),
                        "提交代码commit_id": "get_commit_id", "测试数据集": "test_pic_set", "测试结果": "result_staus"}
    # 总表数据
    table_td = html.TABLE_TMPL_TOTAL % dict(commit_times_id=test_result_data["commit_times_id"],
                                            测试时间=test_result_data['测试时间'],
                                            提交代码commit_id=test_result_data['提交代码commit_id'],
                                            测试数据集=test_result_data['测试数据集'], 测试结果=test_result_data['测试结果'])
    table_tr0 += table_td
    commit_times_id += 1
    reslut_set_id = 1
    # 详情数据
    test_cases = {"reslut_set_id": reslut_set_id, "测试时间": s["date_time"], "算法版本": s["version"],
                  "测试图片集": s["dataset_path"], "解码率": decoding_rate, "正确率": decoding_accuracy, "耗时": "run_time_avg",
                  "最大耗时": "run_time_max", "最小耗时": "run_time_min", "扫码失败总数": succeed_false_count, "扫码失败文件分布": "扫码失败文件分布"}
    table_td_module = html.TABLE_TMPL_MODULE % dict(reslut_set_id=test_cases["reslut_set_id"], 测试时间=test_cases["测试时间"],
                                                    算法版本=test_cases["算法版本"], 测试图片集=test_cases["测试图片集"],
                                                    解码率=test_cases["解码率"], 正确率=test_cases["正确率"], 耗时=test_cases["耗时"],
                                                    最大耗时=test_cases["最大耗时"], 最小耗时=test_cases["最小耗时"],
                                                    扫码失败总数=test_cases["扫码失败总数"], 扫码失败文件分布=test_cases["扫码失败文件分布"])
    table_tr1 += table_td_module
    # 表头总数
    total_str = '共 %s，通过 %s，失败 %s' % (numfail + numsucc, numsucc, numfail)
    # case数据
    # test_data={"image_path":str(s["decode_results"][i]["image_path"]),	"image_name":str(s["decode_results"][i]["image_path"]).split("/")[-1],	"succeed":s["decode_results"][i]["succeed"],	"run_time"	"type"	"result"	"target_result"	"rect_points"}
    id = 1
    for i in range(len(s["decode_results"])):
        test_data = {"id": id, "image_path": str(s["decode_results"][i]["image_path"]),
                     "image_name": str(s["decode_results"][i]["image_path"]).split("/")[-1],
                     "succeed": s["decode_results"][i]["succeed"], "run_time": "run_time", "type": "type",
                     "result": "result", "target_result": "target_result", "rect_points": "rect_points"}

        table_td_case = html.TABLE_TMPL_CASE % dict(id=test_data["id"], image_path=test_data["image_path"], succeed=test_data["succeed"],
                                                    run_time=test_data["run_time"], type=test_data["type"],
                                                    result=test_data["result"],
                                                    target_result=test_data["target_result"],
                                                    rect_points=test_data["rect_points"])
        table_tr2 += table_td_case
        id += 1
        i += 1
    output = html.HTML_TMPL % dict(value=total_str, table_tr=table_tr1, table_tr2=table_tr0, table_tr3=table_tr2)
    # 生成html报告
    filename = '{date}_TestReport.html'.format(date=time.strftime('%Y%m%d%H%M%S'))
    print(filename)
    # 获取report的路径
    # dir= os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),'report')
    dir = ('')
    filename = os.path.join(dir, filename)
    with open(filename, 'wb') as f:
        f.write(output.encode('utf8'))
