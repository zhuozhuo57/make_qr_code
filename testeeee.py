#
# Image[0] : /mnt/123/DM//类型/打印码/20210713144938690.png
# duration : 48.124 ms
# result : !@#$%^&*()_+-=[]\
# code_type : 1
import configparser
import os

import openpyxl

file = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
print(config)
s = config.read(file + r'\config.ini')
print(s)
excelPath = config.get("Local", "ResultPath")
filePath = config.get("Local", "logPath2")
data = openpyxl.load_workbook(excelPath)
sheet = data['Sheet1']
sheet.cell(1, 1).value = 'PicName'
sheet.cell(1, 2).value = 'mark'
sheet.cell(1, 3).value = "DecodeResult"
sheet.cell(1, 4).value = "code_type"
# def get_txt(path):
#     with open(path,'r',encoding='utf-8')as f:
#         line=f.readlines(300)
#         print(line)
#         print(type(line))
#         for str in line:
#             str1=str.split('\n')
#             print(str1)
#             str2=str1[0]
#             print(str2)
#             str3=(str2.split(':')[0])#,str2.split(':')[1])
#
#             print(str3)
#             print(type(str3))
#             str4 = (str2.split(':')[-1])
#             print(str4)
#             print(type(str4))
str_ = '"-----------------------------------------------'


def json_codes():
    with open(filePath, 'r', encoding="utf-8")as f:
        m = f.read(1000)
        print(m)
        i = 0
        Image = ''
        duration = ''
        result = ''
        code_type = ''
        picname = Image[i]
        if "Image[i]"


if __name__ == "__main__":
    # get_txt(filePath)
    json_codes()
