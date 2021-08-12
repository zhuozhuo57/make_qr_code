import configparser
import os
import re

import openpyxl


file = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
print(config)
s = config.read(file + r'\config.ini')
excelPath = config.get("Local", "ResultPath")
filePath = config.get("Local", "logPath")
# tempfile = config.get("Local", "temp")
data = openpyxl.load_workbook(excelPath)
sheet = data['Sheet1']
sheet.cell(1, 1).value = 'date_time'
sheet.cell(1, 2).value = "version"
sheet.cell(1, 3).value = "dataset_path"
count = 0
succeed_false_count = 0
succeed_true_count = 0
image_path_line_count=0
# sheet.cell(1, 4).value = "code_type"
with open(filePath, 'r', encoding="utf-8") as f:
    readlines = f.readlines()
    print("readlines  is ", readlines)
    i = 0
    for s in readlines:

        if ('date_time' in s):
            # print("date_time is",s)
            # date_time = re.findall(r'"date_time":(.*)"version":(.*),"dataset_path":(.*)',s)
            date_time = re.findall(r'"date_time":(.*),"version":(.*),"dataset_path":(.*)', s)
            # print("date_time is" ,date_time)
            # print("date_time type  is" ,type(date_time))
            # date_time[0]
            # 实现提取date_time
            # print(date_time[0])
            # print(type(date_time[0]))
            # 实现date_time_value 写入
            date_time_value = date_time[0][0]
            # print("date_time_value is",date_time_value)
            # print(type((date_time[0])[0]))
            sheet.cell(2, 1).value = date_time_value
            data.save(excelPath)
            # 实现version写入
            version = date_time[0][1]
            # print("version_value is",version)
            # print(type((date_time[0])[1]))
            sheet.cell(2, 2).value = version
            data.save(excelPath)
            # 实现dataset_path写入
            dataset_path = date_time[0][2]
            # ("dataset_path is",dataset_path)
            # print(type((date_time[0])[2]))
            sheet.cell(2, 3).value = dataset_path
            data.save(excelPath)
        # 实现decode_results 写入
        if ('decode_results' in s):
            sheet.cell(3, 1).value = 'decode_results'
            data.save(excelPath)
        if ('image_path' in s):
            sheet.cell(4, 1).value = 'image_path'
            data.save(excelPath)
        if ('succeed' in s):
            sheet.cell(4, 2).value = 'succeed'
            data.save(excelPath)
        if ('run_time' in s):
            sheet.cell(4, 3).value = 'run_time'
            data.save(excelPath)
        if ('results' in s):
            sheet.cell(4, 4).value = 'results'
            data.save(excelPath)

        if ('image_path' in s):
            # 实现image_name写入表中
            image_path_line = re.findall(r'"image_path":(.*),"succeed":(.*),"run_time":(.*),"results":(.*)', s)
            # print("image_path_line is", image_path_line)
            # print("succeed is", s[1])
            # print("image_path_line type  is", type(image_path_line))
            image_path_line_count+=1
            print("image_path_line_count is",image_path_line_count)
            image_name = image_path_line[0]
            #print("image_path_line is", image_path_line.pop(0)[0][1:-1])
            # print("image_name is", str(image_name).strip('""').split('"',1))
            # print("image_name type is", type(str(image_name)))
            image_name_value=image_path_line.pop(0)[0][1:-1]
            print("image_name_value is",image_name_value)
            print("image_name_value type is",type(image_name_value))
            # print("image_name_value is ",list(image_name_value).pop(-2)[1:-1])
            # print("image_name_value is",list(image_name_value)[1:-1])
            sheet.cell(i + 5, 1).value = image_name_value
            data.save(excelPath)
            # 实现succed写入表中
            succeed = list(image_name).pop(1)[1:-1]
            print("succeed is", (succeed))
            if (succeed == 'false'):
                succeed_false_count += 1
                print("succeed_false_count is ", succeed_false_count)
            elif (succeed == 'true'):
                succeed_true_count += 1
                print("succeed_true_count is ", succeed_true_count)

            print(succeed_false_count, succeed_true_count)
            print(type(succeed))
            sheet.cell(i + 5, 2).value = succeed
            data.save(excelPath)
            # 实现runtime写入表中
            run_time = image_name[2]
            print("run_time is", run_time)
            print(type(run_time))
            sheet.cell(i + 5, 3).value = run_time
            data.save(excelPath)
            # 实现results写入表中
            results = image_name[3][1:-3]
            print("results is", results)
            # print("results is", results)
            #print("results is", list(results).pop(0)[0][1:-1])
            #results_value=list(results).pop(0)[1:-1][1:-1]
            # result_value=list(results).remove('}')
            print("results is", results)
            # print("results_value is", results_value)
            # print(type(results_value))
            sheet.cell(i + 5, 4).value = results
            data.save(excelPath)
            i += 1
            print("i number is", i)

       # 实现统计行数及结果判断得出解码率
            total=["统计","图片数量（数据集所有图片）","解码数量（读取成功图片）","解码正确图片","解码失败图片","解码率（解码数量除以图片数量）","解码正确率（解码正确图除以解码数量）","最大解码耗时","最小解码耗时","平均解码耗时","解码失败类型/数量"]
            j=0
            for j in range(len(total)-1):
                sheet.cell(j+4, 5).value = total[j]
                print("total is",total[j])
                # sheet.cell(4, 6).value = '图片数量（数据集所有图片）'
                # sheet.cell(4, 7).value = '解码数量（读取成功图片）'
                # sheet.cell(4, 6).value = '解码正确图片'
                # sheet.cell(4, 6).value = '解码失败图片'
                # sheet.cell(4, 6).value = '解码率（解码数量除以图片数量）'
                # sheet.cell(4, 6).value = '解码正确率（解码正确图除以解码数量）'
                # sheet.cell(4, 6).value = '最大解码耗时'
                # sheet.cell(4, 6).value = '最小解码耗时'
                # sheet.cell(4, 6).value = '平均解码耗时'
                # sheet.cell(4, 6).value = '解码失败类型/数量'
                j+=1
            data.save(excelPath)
            # total_results=[""]
            # x="正确比对结果数量计算"
            # '''最大耗时'''
            # run_time_vlaues=[]
            # h=0
            # mmm=join(run_time)
            # h+=1
            # print("run_time_vlaues is ",str(mmm))
            # run_time_max=run_time.max
            # run_time_mix=run_time.mix
            # run_time_avg=run_time.avg
            # with open('','r',encoding="utf-8")as f:
            #     total_results[1] =len(f.readlines())
            # total_results[2]=image_path_line_count
            # total_results[3]=succeed_true_count
            # total_results[4]=succeed_false_count
            # total_results[5]=total_results[2]/total_results[1]
            # total_results[6]=x/succeed_true_count
            # total_results[7]=x/succeed_true_count
            #
            # if (succeed=="false"):
            #     succeed_false_count=i-1
            #     #count+=1
            #     print("succeed_false_count is",succeed_false_count)
            #     #count+=1
            #     #print("count is ",count,succeed_false_count)
            # else:
            #     succeed_true_count = i+1
            #     #count+=1
            #     print("succeed_true_count is", succeed_true_count)
            #     #count+=1
            #
            # print("succeed_false_count is", succeed_false_count)
            # print("succeed_true_count is", succeed_true_count)
            # print("count is ", succeed_true_count+ succeed_false_count)

'''            if(succeed=='false'):
                succeed_false_count+=1
                print("succeed_false_count is ", succeed_false_count)
            elif(succeed=='true'):
                succeed_true_count+=1
                print("succeed_true_count is ", succeed_true_count)
            print(succeed_false_count,succeed_true_count)'''