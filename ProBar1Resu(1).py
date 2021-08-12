import configparser
import os
import re

import openpyxl

file = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
print(config)
s = config.read(file + r'\config.ini')
print(s)
excelPath = config.get("Local", "ResultPath")
# excelPath ='F:\\备份文件\\杨佳脚本\\test.xlsx'
print('excelPath is', excelPath)
filePath = config.get("Local", "logPath")
# filePath = 'F:\\备份文件\\杨佳脚本\\20210727barcoderesult.txt'
print('filePath is', filePath)

with open(filePath, 'r+', encoding="utf-8")as file:
    print("file i", file)
    allLines = file.readlines()
    file.seek(0)
    count = 0
    for line in allLines:
        if (': ,' in line):
            line = line.replace(': ,', 'null')
    file.write(line)
    file.close()
    print('count is', count)
data = openpyxl.load_workbook(excelPath)
sheet = data['Sheet1']
sheet.cell(1, 1).value = 'PicName'
sheet.cell(1, 2).value = "DecodeResult"
sheet.cell(1, 3).value = "processTime"
with open(filePath, 'r', encoding="utf-8") as raw:
    refile = raw.read()
    result = re.findall(r'%%%(.*)%%%', refile)
    print(result)
    print(len(result))
    picName = re.findall(r'"(.*)png" :', refile)
    print(picName)
    print(len(picName))
    ProcssTime = re.findall(r'\$\$\$(.*)\$\$\$', refile)
    print(ProcssTime)
    print(len(ProcssTime))
    # rate = 1 - count / len(result)
    # if (len(picName) == len(result) & len(ProcssTime) == len(picName)):
    #     print('the process result is OK, decode rate is $rate', rate)

for row in range(2, len(result) + 2):  # print(row)
    sheet.cell(row, 2).value = result[row - 2]
for row in range(2, len(picName) + 2):
    sheet.cell(row, 1).value = picName[row - 2]
for row in range(2, len(ProcssTime) + 2):
    sheet.cell(row, 3).value = ProcssTime[row - 2]
data.save(excelPath)
