# -*- coding:utf-8 -*-

import configparser
import os
import re

import openpyxl

file = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
print(config)
s = config.read(file + r'\config.ini')
print(s)
excelPath = config.get("Local", "ResultPath")
filePath = config.get("Local", "logPath")
data = openpyxl.load_workbook(excelPath)
sheet = data['Sheet1']
sheet.cell(1, 1).value = 'PicName'
sheet.cell(1, 2).value = 'mark'
sheet.cell(1, 3).value = "DecodeResult"
sheet.cell(1, 4).value = "code_type"

file = open(filePath, 'r+', encoding="utf-8")
allLines = file.readlines()
file.seek(0)
count = 0
for line in allLines:
    if ('"",' in line):
        line1 = line.replace('"",', '"[]%%%null%%%&&&null&&&",')
        file.write(line1)
for line2 in allLines:
    if (': "[]' in line2):
        line3 = line2.replace(': "[]', ': "[null]')
        file.write(line3)
if ('null' in line):
    count += 1
file.close()
print(count)
with open(filePath, 'r', encoding="utf-8") as raw:
    refile = raw.read()
    # add [] result
    mark = re.findall(r'\[(.*)\]', refile)
    print("mark", mark)
    print(len(mark))
    # duration: 358.177ms
    result = re.findall(r'duration : (.?) ms"', refile)
    print("result is", result)
    print(len(result))
    # Image[2011] : /mnt/DM//材质/条码纸/20210713140932350.png
    picName = re.findall(r'"(.*).png"', refile)
    print(picName)
    print(len(picName))
    code_type = re.findall(r'\&\&\&(.*)\&\&\&', refile)
    print(code_type)
    print(len(code_type))
for row in range(2, len(picName) + 1):
    sheet.cell(row, 1).value = picName[row - 1]
for row in range(2, len(mark) + 1):
    sheet.cell(row, 2).value = mark[row - 1]
for row in range(2, len(result) + 1):
    sheet.cell(row, 3).value = result[row - 1]
for row in range(2, len(code_type) + 1):
    sheet.cell(row, 4).value = code_type[row - 1]
data.save(excelPath)
